from __future__ import annotations

import argparse
from datetime import datetime
from html import escape
import os
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_REPORT_TITLE = "Nakama Sales Report"
DEFAULT_INPUT_FOLDER = ROOT
DEFAULT_SALES_FILE = DEFAULT_INPUT_FOLDER
DEFAULT_PRODUCTS_FILE = DEFAULT_INPUT_FOLDER
DEFAULT_HTML_OUTPUT_FILE = Path(os.environ.get("SALES_REPORT_HTML_OUTPUT", ROOT / f"{DEFAULT_REPORT_TITLE}.html"))
DEFAULT_EXCEL_OUTPUT_FILE = Path(os.environ.get("SALES_REPORT_EXCEL_OUTPUT", ROOT / f"{DEFAULT_REPORT_TITLE}.xlsx"))
SALES_FILE = DEFAULT_SALES_FILE
PRODUCTS_FILE = DEFAULT_PRODUCTS_FILE
EXCLUDED_CUSTOMERS = {"\u5927\u7edf\u534e\u7279\u4ef7\u6d3b\u52a8\u7559\u8d27"}


def money(value: float) -> str:
    return f"${value:,.2f}"


def number(value: float) -> str:
    return f"{value:,.0f}"


def pct(value: float) -> str:
    return f"{value:.1%}"


def price_label(min_price: float, max_price: float) -> str:
    if min_price == max_price:
        return money(min_price)
    return f"{money(min_price)} - {money(max_price)}"


def nonzero_min(series: pd.Series) -> float:
    nonzero = series[series > 0]
    if nonzero.empty:
        return 0
    return float(nonzero.min())


def nonzero_max(series: pd.Series) -> float:
    nonzero = series[series > 0]
    if nonzero.empty:
        return 0
    return float(nonzero.max())


def text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="EEF2F4")
    for cell in row:
        cell.fill = fill
        cell.font = Font(bold=True, color="3C4852")
        cell.alignment = Alignment(vertical="center")


def set_widths(sheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def add_bar_labels(chart: BarChart) -> None:
    chart.legend = None
    chart.dLbls = DataLabelList()
    chart.dLbls.showCatName = True
    chart.dLbls.showVal = True
    chart.dLbls.showLegendKey = False
    chart.dLbls.showSerName = False
    chart.y_axis.delete = False
    chart.y_axis.tickLblPos = "nextTo"


def write_excel_report(
    grouped: pd.DataFrame,
    product_grouped: pd.DataFrame,
    customer_grouped: pd.DataFrame,
    customer_detail_grouped: pd.DataFrame,
    customer_brand_matrix: pd.DataFrame,
    product_monthly_grouped: pd.DataFrame,
    product_monthly_matrix: pd.DataFrame,
    total_revenue: float,
    excel_output_file: Path,
) -> Path:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    details = workbook.create_sheet("Brand Details")
    charts = workbook.create_sheet("Charts", 0)
    customer_summary = workbook.create_sheet("Customer Summary")
    customer_details = workbook.create_sheet("Customer Details")
    customer_matrix = workbook.create_sheet("Customer Brand Matrix")
    top_products = workbook.create_sheet("Top Products")
    monthly_top_products = workbook.create_sheet("Monthly Top Products")
    product_monthly_matrix_sheet = workbook.create_sheet("Product Monthly Matrix")

    summary.append(["Brand", "Units Sold", "Sales", "Sold SKUs", "Sales Share"])
    style_header(summary[1])
    for _, row in grouped.iterrows():
        share = row["Revenue"] / total_revenue if total_revenue else 0
        summary.append([str(row["Brand"]), row["Quantity"], row["Revenue"], int(row["SkuCount"]), share])

    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions
    set_widths(summary, [28, 14, 16, 12, 12])
    for row in summary.iter_rows(min_row=2, min_col=2, max_col=5):
        row[0].number_format = '#,##0'
        row[1].number_format = '$#,##0.00'
        row[2].number_format = '0'
        row[3].number_format = '0.0%'

    top_count = min(10, len(grouped))
    charts["A1"] = "Top Brand Sales"
    charts["A1"].font = Font(bold=True, size=16)
    chart_data = [["Brand", "Sales", "Units Sold", "Sales Share"]]
    for _, row in grouped.head(top_count).iterrows():
        share = row["Revenue"] / total_revenue if total_revenue else 0
        chart_data.append([str(row["Brand"]), row["Revenue"], row["Quantity"], share])
    for row in chart_data:
        charts.append(row)
    style_header(charts[3])
    set_widths(charts, [26, 16, 14, 14])
    for row in charts.iter_rows(min_row=4, max_row=top_count + 3, min_col=2, max_col=4):
        row[0].number_format = '$#,##0.00'
        row[1].number_format = '#,##0'
        row[2].number_format = '0.0%'

    sales_chart = BarChart()
    sales_chart.type = "bar"
    sales_chart.style = 10
    sales_chart.title = "Top Brand Sales"
    sales_chart.y_axis.title = "Brand"
    sales_chart.x_axis.title = "Sales"
    sales_chart.height = 8
    sales_chart.width = 18
    sales_chart.add_data(Reference(charts, min_col=2, min_row=3, max_row=top_count + 3), titles_from_data=True)
    sales_chart.set_categories(Reference(charts, min_col=1, min_row=4, max_row=top_count + 3))
    add_bar_labels(sales_chart)
    charts.add_chart(sales_chart, "F2")

    units_chart = BarChart()
    units_chart.type = "bar"
    units_chart.style = 11
    units_chart.title = "Top Brand Units Sold"
    units_chart.y_axis.title = "Brand"
    units_chart.x_axis.title = "Units"
    units_chart.height = 8
    units_chart.width = 18
    units_chart.add_data(Reference(charts, min_col=3, min_row=3, max_row=top_count + 3), titles_from_data=True)
    units_chart.set_categories(Reference(charts, min_col=1, min_row=4, max_row=top_count + 3))
    add_bar_labels(units_chart)
    charts.add_chart(units_chart, "F20")

    share_chart = PieChart()
    share_chart.title = "Top Brand Sales Share"
    share_chart.height = 8
    share_chart.width = 12
    share_chart.add_data(Reference(charts, min_col=4, min_row=3, max_row=top_count + 3), titles_from_data=True)
    share_chart.set_categories(Reference(charts, min_col=1, min_row=4, max_row=top_count + 3))
    charts.add_chart(share_chart, "X2")

    details.sheet_properties.outlinePr.summaryBelow = False
    details.append(["Brand/Product", "SKU", "Barcode", "Units Sold", "Sales", "Price"])
    style_header(details[1])
    details.freeze_panes = "A2"
    set_widths(details, [52, 18, 22, 14, 16, 20])

    brand_fill = PatternFill("solid", fgColor="F7F8FA")
    for _, brand_row in grouped.iterrows():
        brand = str(brand_row["Brand"])
        share = brand_row["Revenue"] / total_revenue if total_revenue else 0
        details.append(
            [
                brand,
                "",
                "",
                brand_row["Quantity"],
                brand_row["Revenue"],
                f"{int(brand_row['SkuCount'])} SKUs / {pct(share)}",
            ]
        )
        brand_excel_row = details.max_row
        for cell in details[brand_excel_row]:
            cell.fill = brand_fill
            cell.font = Font(bold=True)
        details.cell(brand_excel_row, 4).number_format = '#,##0'
        details.cell(brand_excel_row, 5).number_format = '$#,##0.00'

        products = product_grouped[product_grouped["Brand"].eq(brand)]
        for _, product in products.iterrows():
            details.append(
                [
                    text(product["Product Name"]),
                    text(product["SKU"]),
                    text(product["Barcode Text"]),
                    product["Quantity"],
                    product["Revenue"],
                    price_label(product["MinPrice"], product["MaxPrice"]),
                ]
            )
            product_excel_row = details.max_row
            details.row_dimensions[product_excel_row].outlineLevel = 1
            details.row_dimensions[product_excel_row].hidden = True
            details.cell(product_excel_row, 4).number_format = '#,##0'
            details.cell(product_excel_row, 5).number_format = '$#,##0.00'

    details.auto_filter.ref = details.dimensions
    customer_summary.append(
        ["Customer", "Sales", "Units Sold", "Orders", "Sold SKUs", "Top Brand", "Top Product", "Last Order Date"]
    )
    style_header(customer_summary[1])
    for _, row in customer_grouped.iterrows():
        customer_summary.append(
            [
                str(row["Customer Name"]),
                row["Revenue"],
                row["Quantity"],
                int(row["Orders"]),
                int(row["SkuCount"]),
                str(row["TopBrand"]),
                str(row["TopProduct"]),
                row["LastOrderDate"],
            ]
        )
    customer_summary.freeze_panes = "A2"
    customer_summary.auto_filter.ref = customer_summary.dimensions
    set_widths(customer_summary, [48, 16, 14, 10, 12, 22, 42, 16])
    for row in customer_summary.iter_rows(min_row=2, min_col=2, max_col=5):
        row[0].number_format = '$#,##0.00'
        row[1].number_format = '#,##0'
        row[2].number_format = '0'
        row[3].number_format = '0'
    for cell in customer_summary["H"][1:]:
        cell.number_format = "yyyy-mm-dd"

    customer_details.sheet_properties.outlinePr.summaryBelow = False
    customer_details.append(["Customer/Brand/Product", "Brand", "Product", "SKU", "Barcode", "Units Sold", "Sales", "Price"])
    style_header(customer_details[1])
    customer_details.freeze_panes = "A2"
    set_widths(customer_details, [46, 22, 42, 18, 22, 14, 16, 20])
    customer_fill = PatternFill("solid", fgColor="F7F8FA")
    for _, customer_row in customer_grouped.iterrows():
        customer = str(customer_row["Customer Name"])
        customer_details.append(
            [
                customer,
                "",
                "",
                "",
                "",
                customer_row["Quantity"],
                customer_row["Revenue"],
                f"{int(customer_row['Orders'])} orders / {int(customer_row['SkuCount'])} SKUs",
            ]
        )
        customer_excel_row = customer_details.max_row
        for cell in customer_details[customer_excel_row]:
            cell.fill = customer_fill
            cell.font = Font(bold=True)
        customer_details.cell(customer_excel_row, 6).number_format = '#,##0'
        customer_details.cell(customer_excel_row, 7).number_format = '$#,##0.00'

        rows = customer_detail_grouped[customer_detail_grouped["Customer Name"].eq(customer)]
        for _, detail in rows.iterrows():
            customer_details.append(
                [
                    text(detail["Product Name"]),
                    text(detail["Brand"]),
                    text(detail["Product Name"]),
                    text(detail["SKU"]),
                    text(detail["Barcode Text"]),
                    detail["Quantity"],
                    detail["Revenue"],
                    price_label(detail["MinPrice"], detail["MaxPrice"]),
                ]
            )
            detail_row = customer_details.max_row
            customer_details.row_dimensions[detail_row].outlineLevel = 1
            customer_details.row_dimensions[detail_row].hidden = True
            customer_details.cell(detail_row, 6).number_format = '#,##0'
            customer_details.cell(detail_row, 7).number_format = '$#,##0.00'
    customer_details.auto_filter.ref = customer_details.dimensions

    customer_matrix.append(list(customer_brand_matrix.columns))
    style_header(customer_matrix[1])
    for row in customer_brand_matrix.itertuples(index=False):
        customer_matrix.append(list(row))
    customer_matrix.freeze_panes = "B2"
    customer_matrix.auto_filter.ref = customer_matrix.dimensions
    customer_matrix.column_dimensions["A"].width = 48
    for col in range(2, customer_matrix.max_column + 1):
        customer_matrix.column_dimensions[get_column_letter(col)].width = 14
    for row in customer_matrix.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = '$#,##0'

    top_products.append(["Rank", "Brand", "Product", "SKU", "Barcode", "Units Sold", "Sales", "Price"])
    style_header(top_products[1])
    set_widths(top_products, [8, 22, 48, 18, 24, 14, 16, 20])
    for rank, (_, row) in enumerate(product_grouped.sort_values("Quantity", ascending=False).head(50).iterrows(), start=1):
        top_products.append(
            [
                rank,
                text(row["Brand"]),
                text(row["Product Name"]),
                text(row["SKU"]),
                text(row["Barcode Text"]),
                row["Quantity"],
                row["Revenue"],
                price_label(row["MinPrice"], row["MaxPrice"]),
            ]
        )
    top_products.freeze_panes = "A2"
    top_products.auto_filter.ref = top_products.dimensions
    for row in top_products.iter_rows(min_row=2, min_col=6, max_col=7):
        row[0].number_format = "#,##0"
        row[1].number_format = "$#,##0.00"

    monthly_top_products.append(["Month", "Rank", "Brand", "Product", "SKU", "Barcode", "Units Sold", "Sales"])
    style_header(monthly_top_products[1])
    set_widths(monthly_top_products, [14, 8, 22, 48, 18, 24, 14, 16])
    for month in sorted(product_monthly_grouped["Order Month"].astype(str).unique()):
        month_rows = product_monthly_grouped[product_monthly_grouped["Order Month"].astype(str).eq(month)]
        for rank, (_, row) in enumerate(month_rows.sort_values("Quantity", ascending=False).head(50).iterrows(), start=1):
            monthly_top_products.append(
                [
                    month,
                    rank,
                    text(row["Brand"]),
                    text(row["Product Name"]),
                    text(row["SKU"]),
                    text(row["Barcode Text"]),
                    row["Quantity"],
                    row["Revenue"],
                ]
            )
    monthly_top_products.freeze_panes = "A2"
    monthly_top_products.auto_filter.ref = monthly_top_products.dimensions
    for row in monthly_top_products.iter_rows(min_row=2, min_col=7, max_col=8):
        row[0].number_format = "#,##0"
        row[1].number_format = "$#,##0.00"

    product_monthly_matrix_sheet.append(list(product_monthly_matrix.columns))
    style_header(product_monthly_matrix_sheet[1])
    for row in product_monthly_matrix.itertuples(index=False):
        product_monthly_matrix_sheet.append(list(row))
    product_monthly_matrix_sheet.freeze_panes = "E2"
    product_monthly_matrix_sheet.auto_filter.ref = product_monthly_matrix_sheet.dimensions

    month_columns = max(0, product_monthly_matrix.shape[1] - 4)
    set_widths(
        product_monthly_matrix_sheet,
        [22, 48, 18, 24] + [12] * month_columns,
    )
    for row in product_monthly_matrix_sheet.iter_rows(min_row=2, min_col=5):
        for cell in row:
            cell.number_format = "#,##0"

    customer_top_count = min(10, len(customer_grouped))
    customer_chart_start = 16
    charts.cell(customer_chart_start, 1, "Customer")
    charts.cell(customer_chart_start, 2, "Sales")
    charts.cell(customer_chart_start, 3, "Units Sold")
    style_header(charts[customer_chart_start])
    for idx, (_, row) in enumerate(customer_grouped.head(customer_top_count).iterrows(), start=customer_chart_start + 1):
        charts.cell(idx, 1, str(row["Customer Name"]))
        charts.cell(idx, 2, row["Revenue"])
        charts.cell(idx, 3, row["Quantity"])
        charts.cell(idx, 2).number_format = '$#,##0.00'
        charts.cell(idx, 3).number_format = '#,##0'

    customer_sales_chart = BarChart()
    customer_sales_chart.type = "bar"
    customer_sales_chart.style = 12
    customer_sales_chart.title = "Top Customer Sales"
    customer_sales_chart.y_axis.title = "Customer"
    customer_sales_chart.x_axis.title = "Sales"
    customer_sales_chart.height = 9
    customer_sales_chart.width = 24
    customer_sales_chart.legend = None
    customer_sales_chart.add_data(
        Reference(charts, min_col=2, min_row=customer_chart_start, max_row=customer_chart_start + customer_top_count),
        titles_from_data=True,
    )
    customer_sales_chart.set_categories(
        Reference(charts, min_col=1, min_row=customer_chart_start + 1, max_row=customer_chart_start + customer_top_count)
    )
    add_bar_labels(customer_sales_chart)
    charts.add_chart(customer_sales_chart, "X20")

    customer_units_chart = BarChart()
    customer_units_chart.type = "bar"
    customer_units_chart.style = 13
    customer_units_chart.title = "Top Customer Units Sold"
    customer_units_chart.y_axis.title = "Customer"
    customer_units_chart.x_axis.title = "Units"
    customer_units_chart.height = 9
    customer_units_chart.width = 24
    customer_units_chart.legend = None
    customer_units_chart.add_data(
        Reference(charts, min_col=3, min_row=customer_chart_start, max_row=customer_chart_start + customer_top_count),
        titles_from_data=True,
    )
    customer_units_chart.set_categories(
        Reference(charts, min_col=1, min_row=customer_chart_start + 1, max_row=customer_chart_start + customer_top_count)
    )
    add_bar_labels(customer_units_chart)
    charts.add_chart(customer_units_chart, "X38")

    try:
        workbook.save(excel_output_file)
        return excel_output_file
    except PermissionError:
        fallback = excel_output_file.with_name(
            f"{excel_output_file.stem}-updated-{datetime.now():%Y%m%d-%H%M%S}{excel_output_file.suffix}"
        )
        workbook.save(fallback)
        return fallback


def build_interactive_html_report(
    grouped: pd.DataFrame,
    product_grouped: pd.DataFrame,
    customer_grouped: pd.DataFrame,
    customer_detail_grouped: pd.DataFrame,
    customer_monthly_grouped: pd.DataFrame,
    product_monthly_grouped: pd.DataFrame,
    product_monthly_matrix: pd.DataFrame,
    date_range: str,
    total_quantity: float,
    total_revenue: float,
    total_skus: int,
    report_title: str,
) -> str:
    def records(frame: pd.DataFrame, columns: list[str]) -> list[dict]:
        output = []
        for _, row in frame.iterrows():
            item = {}
            for column in columns:
                value = row[column]
                if pd.isna(value):
                    item[column] = ""
                elif hasattr(value, "strftime"):
                    item[column] = value.strftime("%Y-%m-%d")
                elif isinstance(value, (int, float)):
                    item[column] = float(value)
                else:
                    item[column] = str(value)
            output.append(item)
        return output

    brand_summary = grouped.copy()
    brand_summary["Share"] = brand_summary["Revenue"] / total_revenue if total_revenue else 0
    customer_summary = customer_grouped.copy()
    customer_summary["Share"] = customer_summary["Revenue"] / total_revenue if total_revenue else 0
    top_brand = "" if brand_summary.empty else str(brand_summary.iloc[0]["Brand"])
    top_customer = "" if customer_summary.empty else str(customer_summary.iloc[0]["Customer Name"])
    top_10_brand_share = (
        float(brand_summary.head(10)["Revenue"].sum()) / total_revenue
        if total_revenue and not brand_summary.empty
        else 0
    )

    report_data = {
        "kpis": {
            "totalSales": total_revenue,
            "totalUnits": total_quantity,
            "soldSkus": total_skus,
            "brands": len(grouped),
            "customers": len(customer_grouped),
            "topBrand": top_brand,
            "topCustomer": top_customer,
            "top10BrandShare": top_10_brand_share,
            "dateRange": date_range,
        },
        "brands": records(brand_summary, ["Brand", "Quantity", "Revenue", "SkuCount", "Share"]),
        "products": records(
            product_grouped,
            ["Brand", "Product Name", "SKU", "Barcode Text", "Quantity", "Revenue", "MinPrice", "MaxPrice"],
        ),
        "productMonthlyDetails": records(
            product_monthly_grouped,
            ["Brand", "Product Name", "SKU", "Barcode Text", "Order Month", "Quantity", "Revenue"],
        ),
        "productMonthlyMatrix": records(
            product_monthly_matrix,
            list(product_monthly_matrix.columns),
        ),
        "customers": records(
            customer_summary,
            ["Customer Name", "Revenue", "Quantity", "Orders", "SkuCount", "TopBrand", "TopProduct", "LastOrderDate", "Share"],
        ),
        "customerDetails": records(
            customer_detail_grouped,
            ["Customer Name", "Brand", "Product Name", "SKU", "Barcode Text", "Quantity", "Revenue", "MinPrice", "MaxPrice"],
        ),
        "customerMonthlyDetails": records(
            customer_monthly_grouped,
            ["Customer Name", "Brand", "Product Name", "SKU", "Barcode Text", "Order Month", "Quantity", "Revenue"],
        ),
    }
    data_json = json.dumps(report_data, ensure_ascii=False)

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(report_title)}</title>
  <style>
    :root {{
      --ink: #172026;
      --muted: #65717d;
      --line: #d9dee4;
      --panel: #f7f8fa;
      --accent: #0f766e;
      --accent-2: #3b6ea8;
      --accent-soft: #d8efeb;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, "Microsoft YaHei", sans-serif; color: var(--ink); background: #fff; }}
    header {{ padding: 32px 40px 20px; border-bottom: 1px solid var(--line); }}
    h1 {{ margin: 0 0 8px; font-size: 30px; letter-spacing: 0; }}
    h2 {{ margin: 0 0 14px; font-size: 18px; }}
    h3 {{ margin: 0 0 10px; font-size: 15px; }}
    .subtitle {{ margin: 0; color: var(--muted); }}
    main {{ padding: 24px 40px 42px; }}
    .summary-highlights {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; margin-bottom: 12px; }}
    .highlight {{ border: 1px solid var(--line); border-radius: 8px; padding: 16px; background: #fff; }}
    .highlight .value {{ font-size: 18px; line-height: 1.25; overflow-wrap: anywhere; }}
    .kpis {{ display: grid; grid-template-columns: repeat(5, minmax(140px, 1fr)); gap: 12px; margin-bottom: 24px; }}
    .kpi {{ border: 1px solid var(--line); background: var(--panel); border-radius: 8px; padding: 14px; }}
    .label {{ color: var(--muted); font-size: 12px; text-transform: uppercase; letter-spacing: .05em; }}
    .value {{ margin-top: 5px; font-size: 22px; font-weight: 700; }}
    .tabs {{ display: flex; gap: 8px; margin: 12px 0 22px; border-bottom: 1px solid var(--line); }}
    .tab-button {{ appearance: none; border: 1px solid var(--line); border-bottom: 0; background: #fff; padding: 10px 14px; border-radius: 8px 8px 0 0; cursor: pointer; font-weight: 700; }}
    .tab-button.active {{ color: #fff; background: var(--accent); border-color: var(--accent); }}
    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}
    .toolbar {{ display: flex; flex-wrap: wrap; gap: 10px; align-items: start; margin: 12px 0; }}
    input[type="search"] {{ width: min(420px, 100%); padding: 10px 12px; border: 1px solid var(--line); border-radius: 6px; font-size: 14px; }}
    .filter-control {{ position: relative; width: min(220px, 100%); }}
    .filter-input {{ width: 100%; padding: 10px 12px; border: 1px solid var(--line); border-radius: 6px; font-size: 14px; background: #fff; }}
    select {{ padding: 10px 12px; border: 1px solid var(--line); border-radius: 6px; font-size: 14px; background: #fff; }}
    #productMonthSelect {{ width: 180px; min-width: 140px; }}
    .filter-options {{ display: none; position: absolute; z-index: 20; top: calc(100% + 4px); left: 0; right: 0; max-height: 220px; overflow-y: auto; border: 1px solid var(--line); border-radius: 6px; background: #fff; box-shadow: 0 8px 24px rgba(23,32,38,.12); font-size: 13px; line-height: 1.35; }}
    .filter-options.active {{ display: block; }}
    .filter-option {{ padding: 8px 10px; cursor: pointer; overflow-wrap: anywhere; }}
    .filter-option:hover {{ background: var(--panel); }}
    .section-heading {{ display: flex; align-items: end; justify-content: space-between; gap: 16px; margin: 18px 0 10px; }}
    .section-heading p {{ margin: 0; color: var(--muted); }}
    .charts {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 18px; margin: 18px 0 26px; }}
    .chart-box {{ border: 1px solid var(--line); border-radius: 8px; padding: 16px; min-height: 320px; }}
    .bar-row {{ display: grid; grid-template-columns: minmax(160px, 260px) 1fr minmax(90px, auto); gap: 10px; align-items: center; margin: 8px 0; }}
    .bar-label {{ font-weight: 700; overflow-wrap: anywhere; }}
    .bar-track {{ height: 18px; background: var(--accent-soft); border-radius: 5px; overflow: hidden; }}
    .bar-fill {{ height: 100%; background: var(--accent); }}
    .bar-fill.blue {{ background: var(--accent-2); }}
    .bar-value {{ text-align: right; color: var(--muted); white-space: nowrap; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 9px 8px; text-align: left; vertical-align: top; }}
    th {{ background: #eef2f4; color: #3c4852; font-size: 12px; text-transform: uppercase; letter-spacing: .04em; }}
    .num {{ text-align: right; white-space: nowrap; }}
    details {{ border-bottom: 1px solid var(--line); }}
    summary {{ cursor: pointer; padding: 10px 8px; font-weight: 700; }}
    .detail-wrap {{ padding: 0 0 16px 22px; overflow-x: auto; }}
    .product-month-detail {{ border-bottom: 0; }}
    .product-month-detail summary {{ padding: 0; font-weight: 400; color: var(--accent-2); }}
    .month-detail-table {{ margin-top: 8px; min-width: 360px; }}
    .table-box {{ border: 1px solid var(--line); border-radius: 8px; overflow-x: auto; margin: 18px 0 26px; }}
    @media (max-width: 900px) {{
      header, main {{ padding-left: 18px; padding-right: 18px; }}
      .summary-highlights, .kpis, .charts {{ grid-template-columns: 1fr; }}
      .bar-row {{ grid-template-columns: 1fr; }}
      .bar-value, .num {{ text-align: left; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>{escape(report_title)}</h1>
  </header>
  <main>
    <div class="summary-highlights" id="summaryHighlights"></div>
    <div class="kpis" id="kpis"></div>
    <div class="tabs">
      <button class="tab-button active" data-tab="brand">Brand Dashboard</button>
      <button class="tab-button" data-tab="customer">Customer Dashboard</button>
      <button class="tab-button" data-tab="product">Product Dashboard</button>
    </div>

    <section class="tab-panel active" id="brandPanel">
      <div class="section-heading">
        <div>
          <h2>Brand Performance</h2>
        </div>
      </div>
      <div class="charts">
        <div class="chart-box"><h3>Top Brand Sales</h3><div id="brandSalesChart"></div></div>
        <div class="chart-box"><h3>Top Brand Units Sold</h3><div id="brandUnitsChart"></div></div>
      </div>
      <div class="toolbar">
        <input id="brandSearch" type="search" placeholder="Search brand, product, SKU, barcode">
        <div class="filter-control"><input id="brandBrandFilter" class="filter-input" placeholder="Filter brand"><div class="filter-options" data-for="brandBrandFilter"></div></div>
        <div class="filter-control"><input id="brandProductFilter" class="filter-input" placeholder="Filter product"><div class="filter-options" data-for="brandProductFilter"></div></div>
        <div class="filter-control"><input id="brandSkuFilter" class="filter-input" placeholder="Filter SKU"><div class="filter-options" data-for="brandSkuFilter"></div></div>
        <div class="filter-control"><input id="brandBarcodeFilter" class="filter-input" placeholder="Filter barcode"><div class="filter-options" data-for="brandBarcodeFilter"></div></div>
      </div>
      <div id="brandTable"></div>
    </section>

    <section class="tab-panel" id="customerPanel">
      <div class="section-heading">
        <div>
          <h2>Customer Drill-down</h2>
        </div>
      </div>
      <div class="charts">
        <div class="chart-box"><h3>Top Customer Sales</h3><div id="customerSalesChart"></div></div>
        <div class="chart-box"><h3>Top Customer Units Sold</h3><div id="customerUnitsChart"></div></div>
      </div>
      <div class="toolbar">
        <input id="customerSearch" type="search" placeholder="Search customer, brand, product, SKU, barcode">
        <div class="filter-control"><input id="customerCustomerFilter" class="filter-input" placeholder="Filter customer"><div class="filter-options" data-for="customerCustomerFilter"></div></div>
        <div class="filter-control"><input id="customerBrandFilter" class="filter-input" placeholder="Filter brand"><div class="filter-options" data-for="customerBrandFilter"></div></div>
        <div class="filter-control"><input id="customerProductFilter" class="filter-input" placeholder="Filter product"><div class="filter-options" data-for="customerProductFilter"></div></div>
        <div class="filter-control"><input id="customerSkuFilter" class="filter-input" placeholder="Filter SKU"><div class="filter-options" data-for="customerSkuFilter"></div></div>
        <div class="filter-control"><input id="customerBarcodeFilter" class="filter-input" placeholder="Filter barcode"><div class="filter-options" data-for="customerBarcodeFilter"></div></div>
      </div>
      <div id="customerTable"></div>
    </section>

    <section class="tab-panel" id="productPanel">
      <div class="section-heading">
        <div>
          <h2>Top 50 Products by Units</h2>
        </div>
      </div>
      <div class="toolbar">
        <input id="productSearch" type="search" placeholder="Search brand, product, SKU, barcode">
        <div class="filter-control"><input id="productBrandFilter" class="filter-input" placeholder="Filter brand"><div class="filter-options" data-for="productBrandFilter"></div></div>
        <div class="filter-control"><input id="productProductFilter" class="filter-input" placeholder="Filter product"><div class="filter-options" data-for="productProductFilter"></div></div>
        <div class="filter-control"><input id="productSkuFilter" class="filter-input" placeholder="Filter SKU"><div class="filter-options" data-for="productSkuFilter"></div></div>
        <div class="filter-control"><input id="productBarcodeFilter" class="filter-input" placeholder="Filter barcode"><div class="filter-options" data-for="productBarcodeFilter"></div></div>
      </div>
      <div class="table-box" id="productTotalTable"></div>

      <div class="section-heading">
        <div>
          <h2>Monthly Top 50 Products by Units</h2>
        </div>
        <select id="productMonthSelect"></select>
      </div>
      <div class="table-box" id="productMonthlyTable"></div>

      <div class="section-heading">
        <div>
          <h2>Product Monthly Units Matrix</h2>
        </div>
      </div>
      <div class="table-box" id="productMonthlyMatrixTable"></div>
    </section>

  </main>
  <script>
    window.REPORT_DATA = {data_json};
    const fmtMoney = value => '$' + Number(value || 0).toLocaleString('en-CA', {{minimumFractionDigits: 2, maximumFractionDigits: 2}});
    const fmtNum = value => Number(value || 0).toLocaleString('en-CA', {{maximumFractionDigits: 0}});
    const fmtPct = value => Number(value || 0).toLocaleString('en-CA', {{style: 'percent', minimumFractionDigits: 1, maximumFractionDigits: 1}});
    const esc = value => String(value ?? '').replace(/[&<>"']/g, ch => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[ch]));
    const priceLabel = item => Number(item.MinPrice) === Number(item.MaxPrice) ? fmtMoney(item.MinPrice) : `${{fmtMoney(item.MinPrice)}} - ${{fmtMoney(item.MaxPrice)}}`;
    const norm = value => String(value ?? '').toLowerCase();

    function valueOf(id) {{
      const element = document.getElementById(id);
      return element ? element.value : '';
    }}

    function sortedUnique(rows, key) {{
      return [...new Set(rows.map(row => String(row[key] ?? '').trim()).filter(Boolean))]
        .sort((a, b) => a.localeCompare(b, undefined, {{numeric: true, sensitivity: 'base'}}));
    }}

    function setAutocompleteOptions(id, rows, key) {{
      const input = document.getElementById(id);
      if (!input) return;
      const values = sortedUnique(rows, key);
      input._filterOptions = values;
      renderFilterOptions(input);
    }}

    function renderFilterOptions(input) {{
      const list = document.querySelector(`.filter-options[data-for="${{input.id}}"]`);
      if (!list) return;
      const q = norm(input.value);
      const values = (input._filterOptions || [])
        .filter(value => !q || norm(value).includes(q))
        .slice(0, 80);
      list.innerHTML = values.map(value => `<div class="filter-option" data-value="${{esc(value)}}">${{esc(value)}}</div>`).join('');
      list.classList.toggle('active', document.activeElement === input && values.length > 0);
    }}

    function closeAllFilterOptions(exceptId = '') {{
      document.querySelectorAll('.filter-options').forEach(list => {{
        if (list.dataset.for !== exceptId) list.classList.remove('active');
      }});
    }}

    function setupFilters() {{
      const data = window.REPORT_DATA;
      setAutocompleteOptions('brandBrandFilter', data.brands, 'Brand');
      setAutocompleteOptions('brandProductFilter', data.products, 'Product Name');
      setAutocompleteOptions('brandSkuFilter', data.products, 'SKU');
      setAutocompleteOptions('brandBarcodeFilter', data.products, 'Barcode Text');
      setAutocompleteOptions('customerCustomerFilter', data.customers, 'Customer Name');
      setAutocompleteOptions('customerBrandFilter', data.customerDetails, 'Brand');
      setAutocompleteOptions('customerProductFilter', data.customerDetails, 'Product Name');
      setAutocompleteOptions('customerSkuFilter', data.customerDetails, 'SKU');
      setAutocompleteOptions('customerBarcodeFilter', data.customerDetails, 'Barcode Text');
      setAutocompleteOptions('productBrandFilter', data.products, 'Brand');
      setAutocompleteOptions('productProductFilter', data.products, 'Product Name');
      setAutocompleteOptions('productSkuFilter', data.products, 'SKU');
      setAutocompleteOptions('productBarcodeFilter', data.products, 'Barcode Text');
    }}

    function brandFilters() {{
      return {{
        q: norm(valueOf('brandSearch')),
        brand: norm(valueOf('brandBrandFilter')),
        product: norm(valueOf('brandProductFilter')),
        sku: norm(valueOf('brandSkuFilter')),
        barcode: norm(valueOf('brandBarcodeFilter')),
      }};
    }}

    function customerFilters() {{
      return {{
        q: norm(valueOf('customerSearch')),
        customer: norm(valueOf('customerCustomerFilter')),
        brand: norm(valueOf('customerBrandFilter')),
        product: norm(valueOf('customerProductFilter')),
        sku: norm(valueOf('customerSkuFilter')),
        barcode: norm(valueOf('customerBarcodeFilter')),
      }};
    }}

    function productFilters() {{
      return {{
        q: norm(valueOf('productSearch')),
        brand: norm(valueOf('productBrandFilter')),
        product: norm(valueOf('productProductFilter')),
        sku: norm(valueOf('productSkuFilter')),
        barcode: norm(valueOf('productBarcodeFilter')),
      }};
    }}

    function productMatches(item, filters, includeSearch = true) {{
      if (filters.brand && !norm(item.Brand).includes(filters.brand)) return false;
      if (filters.product && !norm(item['Product Name']).includes(filters.product)) return false;
      if (filters.sku && !norm(item.SKU).includes(filters.sku)) return false;
      if (filters.barcode && !norm(item['Barcode Text']).includes(filters.barcode)) return false;
      if (includeSearch && filters.q && ![item.Brand, item['Product Name'], item.SKU, item['Barcode Text']].some(value => norm(value).includes(filters.q))) return false;
      return true;
    }}

    function customerDetailMatches(item, filters, includeSearch = true) {{
      if (filters.customer && !norm(item['Customer Name']).includes(filters.customer)) return false;
      return productMatches(item, filters, includeSearch);
    }}

    function renderBars(targetId, rows, labelKey, valueKey, colorClass = '') {{
      const target = document.getElementById(targetId);
      const top = rows.slice(0, 10);
      const max = Math.max(...top.map(row => Number(row[valueKey] || 0)), 1);
      target.innerHTML = top.map(row => `
        <div class="bar-row">
          <div class="bar-label">${{esc(row[labelKey])}}</div>
          <div class="bar-track"><div class="bar-fill ${{colorClass}}" style="width:${{Number(row[valueKey] || 0) / max * 100}}%"></div></div>
          <div class="bar-value">${{valueKey === 'Revenue' ? fmtMoney(row[valueKey]) : fmtNum(row[valueKey])}}</div>
        </div>`).join('');
    }}

    function renderBrandTable() {{
      const filters = brandFilters();
      const brands = window.REPORT_DATA.brands.filter(brand => {{
        if (filters.brand && !norm(brand.Brand).includes(filters.brand)) return false;
        const products = window.REPORT_DATA.products.filter(product => product.Brand === brand.Brand && productMatches(product, filters, false));
        const brandSearchMatch = !filters.q || norm(brand.Brand).includes(filters.q);
        return products.length > 0 && (brandSearchMatch || products.some(product => productMatches(product, filters, true)));
      }});
      document.getElementById('brandTable').innerHTML = brands.map(brand => {{
        const brandSearchMatch = !filters.q || norm(brand.Brand).includes(filters.q);
        const products = window.REPORT_DATA.products.filter(product =>
          product.Brand === brand.Brand &&
          productMatches(product, filters, false) &&
          (brandSearchMatch || productMatches(product, filters, true))
        );
        return `<details class="brand-detail">
          <summary>${{esc(brand.Brand)}} | Units: ${{fmtNum(brand.Quantity)}} | Sales: ${{fmtMoney(brand.Revenue)}} | SKUs: ${{fmtNum(brand.SkuCount)}} | Share: ${{fmtPct(brand.Share)}}</summary>
          <div class="detail-wrap"><table>
            <thead><tr><th>Product</th><th>SKU</th><th>Barcode</th><th class="num">Units Sold</th><th class="num">Sales</th><th class="num">Price</th></tr></thead>
            <tbody>${{products.map(product => `<tr><td>${{renderBrandProductMonthlyDetail(product)}}</td><td>${{esc(product.SKU)}}</td><td>${{esc(product['Barcode Text'])}}</td><td class="num">${{fmtNum(product.Quantity)}}</td><td class="num">${{fmtMoney(product.Revenue)}}</td><td class="num">${{priceLabel(product)}}</td></tr>`).join('')}}</tbody>
          </table></div>
        </details>`;
      }}).join('');
    }}

    function brandMonthlyRowsFor(product) {{
      return window.REPORT_DATA.productMonthlyDetails.filter(month =>
        month.Brand === product.Brand &&
        month['Product Name'] === product['Product Name'] &&
        month.SKU === product.SKU &&
        month['Barcode Text'] === product['Barcode Text']
      );
    }}

    function renderBrandProductMonthlyDetail(product) {{
      const rows = brandMonthlyRowsFor(product);
      return `<details class="product-month-detail">
        <summary>${{esc(product['Product Name'])}}</summary>
        <table class="month-detail-table">
          <thead><tr><th>Month</th><th class="num">Units Sold</th><th class="num">Sales</th></tr></thead>
          <tbody>${{rows.map(row => `<tr><td>${{esc(row['Order Month'])}}</td><td class="num">${{fmtNum(row.Quantity)}}</td><td class="num">${{fmtMoney(row.Revenue)}}</td></tr>`).join('')}}</tbody>
        </table>
      </details>`;
    }}

    function renderCustomerTable() {{
      const filters = customerFilters();
      const customers = window.REPORT_DATA.customers.filter(customer => {{
        if (filters.customer && !norm(customer['Customer Name']).includes(filters.customer)) return false;
        const details = window.REPORT_DATA.customerDetails.filter(item => item['Customer Name'] === customer['Customer Name'] && customerDetailMatches(item, filters, false));
        const customerSearchMatch = !filters.q || norm(customer['Customer Name']).includes(filters.q);
        return details.length > 0 && (customerSearchMatch || details.some(item => customerDetailMatches(item, filters, true)));
      }});
      document.getElementById('customerTable').innerHTML = customers.map(customer => {{
        const customerSearchMatch = !filters.q || norm(customer['Customer Name']).includes(filters.q);
        const details = window.REPORT_DATA.customerDetails.filter(item =>
          item['Customer Name'] === customer['Customer Name'] &&
          customerDetailMatches(item, filters, false) &&
          (customerSearchMatch || customerDetailMatches(item, filters, true))
        );
        return `<details class="customer-detail">
          <summary>${{esc(customer['Customer Name'])}} | Sales: ${{fmtMoney(customer.Revenue)}} | Units: ${{fmtNum(customer.Quantity)}} | Orders: ${{fmtNum(customer.Orders)}} | Top Brand: ${{esc(customer.TopBrand)}}</summary>
          <div class="detail-wrap"><table>
            <thead><tr><th>Brand</th><th>Product</th><th>SKU</th><th>Barcode</th><th class="num">Units Sold</th><th class="num">Sales</th><th class="num">Price</th></tr></thead>
            <tbody>${{details.map(item => `<tr><td>${{esc(item.Brand)}}</td><td>${{renderProductMonthlyDetail(item)}}</td><td>${{esc(item.SKU)}}</td><td>${{esc(item['Barcode Text'])}}</td><td class="num">${{fmtNum(item.Quantity)}}</td><td class="num">${{fmtMoney(item.Revenue)}}</td><td class="num">${{priceLabel(item)}}</td></tr>`).join('')}}</tbody>
          </table></div>
        </details>`;
      }}).join('');
    }}

    function monthlyRowsFor(item) {{
      return window.REPORT_DATA.customerMonthlyDetails.filter(month =>
        month['Customer Name'] === item['Customer Name'] &&
        month.Brand === item.Brand &&
        month['Product Name'] === item['Product Name'] &&
        month.SKU === item.SKU &&
        month['Barcode Text'] === item['Barcode Text']
      );
    }}

    function renderProductMonthlyDetail(item) {{
      const rows = monthlyRowsFor(item);
      return `<details class="product-month-detail">
        <summary>${{esc(item['Product Name'])}}</summary>
        <table class="month-detail-table">
          <thead><tr><th>Month</th><th class="num">Units Sold</th><th class="num">Sales</th></tr></thead>
          <tbody>${{rows.map(row => `<tr><td>${{esc(row['Order Month'])}}</td><td class="num">${{fmtNum(row.Quantity)}}</td><td class="num">${{fmtMoney(row.Revenue)}}</td></tr>`).join('')}}</tbody>
        </table>
      </details>`;
    }}

    function renderTopProducts(targetId, rows, includeMonth = false) {{
      const sorted = [...rows].sort((a, b) => Number(b.Quantity || 0) - Number(a.Quantity || 0)).slice(0, 50);
      document.getElementById(targetId).innerHTML = `<table>
        <thead><tr>${{includeMonth ? '<th>Month</th>' : ''}}<th>Brand</th><th>Product</th><th>SKU</th><th>Barcode</th><th class="num">Units Sold</th><th class="num">Sales</th></tr></thead>
        <tbody>${{sorted.map(item => `<tr>${{includeMonth ? `<td>${{esc(item['Order Month'])}}</td>` : ''}}<td>${{esc(item.Brand)}}</td><td>${{esc(item['Product Name'])}}</td><td>${{esc(item.SKU)}}</td><td>${{esc(item['Barcode Text'])}}</td><td class="num">${{fmtNum(item.Quantity)}}</td><td class="num">${{fmtMoney(item.Revenue)}}</td></tr>`).join('')}}</tbody>
      </table>`;
    }}

    function renderProductMonthlyMatrix(rows) {{
      const target = document.getElementById('productMonthlyMatrixTable');
      if (!rows.length) {{
        target.innerHTML = '<table><tbody><tr><td>No matching products.</td></tr></tbody></table>';
        return;
      }}
      const columns = Object.keys(rows[0]);
      const monthColumns = columns.slice(4);
      target.innerHTML = `<table>
        <thead><tr><th>Brand</th><th>Product</th><th>SKU</th><th>Barcode</th>${{monthColumns.map(month => `<th class="num">${{esc(month)}}</th>`).join('')}}</tr></thead>
        <tbody>${{rows.map(item => `<tr><td>${{esc(item.Brand)}}</td><td>${{esc(item['Product Name'])}}</td><td>${{esc(item.SKU)}}</td><td>${{esc(item['Barcode Text'])}}</td>${{monthColumns.map(month => `<td class="num">${{fmtNum(item[month])}}</td>`).join('')}}</tr>`).join('')}}</tbody>
      </table>`;
    }}

    function renderProductDashboard() {{
      const filters = productFilters();
      const productRows = window.REPORT_DATA.products.filter(item => productMatches(item, filters, true));
      renderTopProducts('productTotalTable', productRows, false);
      const months = [...new Set(window.REPORT_DATA.productMonthlyDetails.map(item => item['Order Month']))].sort();
      const select = document.getElementById('productMonthSelect');
      if (!select.dataset.ready) {{
        select.innerHTML = months.map(month => `<option value="${{esc(month)}}">${{esc(month)}}</option>`).join('');
        select.dataset.ready = 'true';
      }}
      const selectedMonth = select.value || months[0] || '';
      const monthlyRows = window.REPORT_DATA.productMonthlyDetails.filter(item => item['Order Month'] === selectedMonth && productMatches(item, filters, true));
      renderTopProducts('productMonthlyTable', monthlyRows, true);
      const matrixRows = window.REPORT_DATA.productMonthlyMatrix.filter(item => productMatches(item, filters, true));
      renderProductMonthlyMatrix(matrixRows);
    }}

    function renderDashboard() {{
      const data = window.REPORT_DATA;
      document.getElementById('summaryHighlights').innerHTML = [
        ['Top Brand', data.kpis.topBrand],
        ['Top Customer', data.kpis.topCustomer],
        ['Top 10 Brand Share', fmtPct(data.kpis.top10BrandShare)]
      ].map(item => `<div class="highlight"><div class="label">${{item[0]}}</div><div class="value">${{esc(item[1])}}</div></div>`).join('');
      document.getElementById('kpis').innerHTML = [
        ['Total Sales', fmtMoney(data.kpis.totalSales)],
        ['Units Sold', fmtNum(data.kpis.totalUnits)],
        ['Sold SKUs', fmtNum(data.kpis.soldSkus)],
        ['Brands', fmtNum(data.kpis.brands)],
        ['Customers', fmtNum(data.kpis.customers)]
      ].map(item => `<div class="kpi"><div class="label">${{item[0]}}</div><div class="value">${{item[1]}}</div></div>`).join('');
      renderBars('brandSalesChart', data.brands, 'Brand', 'Revenue');
      renderBars('brandUnitsChart', data.brands, 'Brand', 'Quantity', 'blue');
      renderBars('customerSalesChart', data.customers, 'Customer Name', 'Revenue');
      renderBars('customerUnitsChart', data.customers, 'Customer Name', 'Quantity', 'blue');
      setupFilters();
      renderBrandTable();
      renderCustomerTable();
      renderProductDashboard();
    }}

    document.querySelectorAll('.tab-button').forEach(button => button.addEventListener('click', () => {{
      document.querySelectorAll('.tab-button').forEach(item => item.classList.remove('active'));
      document.querySelectorAll('.tab-panel').forEach(item => item.classList.remove('active'));
      button.classList.add('active');
      document.getElementById(button.dataset.tab + 'Panel').classList.add('active');
    }}));
    ['brandSearch', 'brandBrandFilter', 'brandProductFilter', 'brandSkuFilter', 'brandBarcodeFilter']
      .forEach(id => document.getElementById(id).addEventListener('input', event => {{
        if (event.target.classList.contains('filter-input')) renderFilterOptions(event.target);
        renderBrandTable();
      }}));
    ['customerSearch', 'customerCustomerFilter', 'customerBrandFilter', 'customerProductFilter', 'customerSkuFilter', 'customerBarcodeFilter']
      .forEach(id => document.getElementById(id).addEventListener('input', event => {{
        if (event.target.classList.contains('filter-input')) renderFilterOptions(event.target);
        renderCustomerTable();
      }}));
    ['productSearch', 'productBrandFilter', 'productProductFilter', 'productSkuFilter', 'productBarcodeFilter']
      .forEach(id => document.getElementById(id).addEventListener('input', event => {{
        if (event.target.classList.contains('filter-input')) renderFilterOptions(event.target);
        renderProductDashboard();
      }}));
    document.querySelectorAll('.filter-input').forEach(input => {{
      input.addEventListener('focus', () => {{
        closeAllFilterOptions(input.id);
        renderFilterOptions(input);
      }});
    }});
    document.addEventListener('mousedown', event => {{
      const option = event.target.closest('.filter-option');
      if (option) {{
        const list = option.closest('.filter-options');
        const input = document.getElementById(list.dataset.for);
        input.value = option.dataset.value;
        list.classList.remove('active');
        input.dispatchEvent(new Event('input', {{ bubbles: true }}));
        return;
      }}
      if (!event.target.closest('.filter-control')) {{
        closeAllFilterOptions();
      }}
    }});
    document.getElementById('productMonthSelect').addEventListener('change', renderProductDashboard);
    renderDashboard();
  </script>
</body>
</html>"""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Nakama sales reports in HTML and Excel formats.")
    parser.add_argument("--sales-file", type=Path, default=DEFAULT_SALES_FILE)
    parser.add_argument("--products-file", type=Path, default=DEFAULT_PRODUCTS_FILE)
    parser.add_argument("--html-output", type=Path, default=DEFAULT_HTML_OUTPUT_FILE)
    parser.add_argument("--excel-output", type=Path, default=DEFAULT_EXCEL_OUTPUT_FILE)
    parser.add_argument("--report-title", default=DEFAULT_REPORT_TITLE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    sales_file = args.sales_file
    products_file = args.products_file
    html_output_file = args.html_output
    excel_output_file = args.excel_output
    report_title = args.report_title

    if not sales_file.is_file():
        raise SystemExit(
            f"Sales Orders Excel file is required. Choose a .xlsx file in the GUI or pass --sales-file. Default folder: {DEFAULT_INPUT_FOLDER}"
        )
    if not products_file.is_file():
        raise SystemExit(
            f"Products Excel file is required. Choose a .xlsx file in the GUI or pass --products-file. Default folder: {DEFAULT_INPUT_FOLDER}"
        )

    sales = pd.read_excel(sales_file, dtype={"SKU": "string"})
    products = pd.read_excel(products_file, dtype={"Variant SKU": "string", "Master SKU": "string"})
    sales["Customer"] = sales["Customer"].fillna("").astype(str).str.strip()
    sales = sales[~sales["Customer"].isin(EXCLUDED_CUSTOMERS)].copy()

    products["Brand"] = products["Brand"].fillna("").astype(str).str.strip().replace("", "Unbranded")
    sku_brand = {}
    for _, row in products.iterrows():
        for sku_col in ("Variant SKU", "Master SKU"):
            sku = text(row.get(sku_col))
            if sku and sku not in sku_brand:
                sku_brand[sku] = row["Brand"]

    sales["SKU"] = sales["SKU"].fillna("").astype(str).str.strip()
    sales["Brand"] = sales["SKU"].map(sku_brand).fillna("Unknown / SKU not in product list")
    sales["Customer Name"] = sales["Customer"].replace("", "Unknown Customer")
    sales["Product Name"] = sales["Product"].fillna("").astype(str).str.strip()
    sales["Barcode Text"] = sales["Barcode"].fillna("").astype(str).str.strip()
    sales["Quantity"] = pd.to_numeric(sales["Quantity"], errors="coerce").fillna(0)
    sales["Price"] = pd.to_numeric(sales["Price"], errors="coerce").fillna(0)
    sales["Line Total"] = pd.to_numeric(sales["Total.1"], errors="coerce").fillna(0)
    sales.loc[(sales["Line Total"] == 0) & (sales["Quantity"] != 0), "Line Total"] = (
        sales["Quantity"] * sales["Price"]
    )

    created_dates = pd.to_datetime(sales["Creation date"], errors="coerce")
    sales["Created Date"] = created_dates
    date_min = created_dates.min()
    date_max = created_dates.max()
    if pd.isna(date_min) or pd.isna(date_max):
        date_range = "From source file"
    else:
        date_range = f"{date_min:%Y-%m-%d} to {date_max:%Y-%m-%d}"

    grouped = (
        sales.groupby("Brand", dropna=False)
        .agg(
            Quantity=("Quantity", "sum"),
            Revenue=("Line Total", "sum"),
            SkuCount=("SKU", lambda s: s[s != ""].nunique()),
            Lines=("SKU", "size"),
        )
        .reset_index()
    )
    grouped = grouped.sort_values("Revenue", ascending=False)

    product_grouped = (
        sales.groupby(["Brand", "Product Name", "SKU", "Barcode Text"], dropna=False)
        .agg(
            Quantity=("Quantity", "sum"),
            Revenue=("Line Total", "sum"),
            MinPrice=("Price", nonzero_min),
            MaxPrice=("Price", nonzero_max),
        )
        .reset_index()
        .sort_values(["Brand", "Revenue"], ascending=[True, False])
    )

    customer_base = (
        sales.groupby("Customer Name", dropna=False)
        .agg(
            Quantity=("Quantity", "sum"),
            Revenue=("Line Total", "sum"),
            Orders=("Sales order #", lambda s: s.astype(str).str.strip().nunique()),
            SkuCount=("SKU", lambda s: s[s != ""].nunique()),
            LastOrderDate=("Created Date", "max"),
        )
        .reset_index()
    )
    top_customer_brand = (
        sales.groupby(["Customer Name", "Brand"], dropna=False)["Line Total"]
        .sum()
        .reset_index()
        .sort_values(["Customer Name", "Line Total"], ascending=[True, False])
        .drop_duplicates("Customer Name")
        .rename(columns={"Brand": "TopBrand"})
    )[["Customer Name", "TopBrand"]]
    top_customer_product = (
        sales.groupby(["Customer Name", "Product Name"], dropna=False)["Line Total"]
        .sum()
        .reset_index()
        .sort_values(["Customer Name", "Line Total"], ascending=[True, False])
        .drop_duplicates("Customer Name")
        .rename(columns={"Product Name": "TopProduct"})
    )[["Customer Name", "TopProduct"]]
    customer_grouped = (
        customer_base.merge(top_customer_brand, on="Customer Name", how="left")
        .merge(top_customer_product, on="Customer Name", how="left")
        .sort_values("Revenue", ascending=False)
    )

    customer_detail_grouped = (
        sales.groupby(["Customer Name", "Brand", "Product Name", "SKU", "Barcode Text"], dropna=False)
        .agg(
            Quantity=("Quantity", "sum"),
            Revenue=("Line Total", "sum"),
            MinPrice=("Price", nonzero_min),
            MaxPrice=("Price", nonzero_max),
        )
        .reset_index()
        .sort_values(["Customer Name", "Revenue"], ascending=[True, False])
    )

    sales["Order Month"] = sales["Created Date"].dt.strftime("%Y-%m").fillna("Unknown")
    customer_monthly_grouped = (
        sales.groupby(["Customer Name", "Brand", "Product Name", "SKU", "Barcode Text", "Order Month"], dropna=False)
        .agg(
            Quantity=("Quantity", "sum"),
            Revenue=("Line Total", "sum"),
        )
        .reset_index()
        .sort_values(["Customer Name", "Brand", "Product Name", "SKU", "Order Month"])
    )
    product_monthly_grouped = (
        sales.groupby(["Brand", "Product Name", "SKU", "Barcode Text", "Order Month"], dropna=False)
        .agg(
            Quantity=("Quantity", "sum"),
            Revenue=("Line Total", "sum"),
        )
        .reset_index()
        .sort_values(["Brand", "Product Name", "SKU", "Order Month"])
    )
    product_monthly_matrix = (
        product_monthly_grouped.pivot_table(
            index=["Brand", "Product Name", "SKU", "Barcode Text"],
            columns="Order Month",
            values="Quantity",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )
    month_columns = sorted(
        [column for column in product_monthly_matrix.columns if column not in ["Brand", "Product Name", "SKU", "Barcode Text"]]
    )
    product_monthly_matrix = product_monthly_matrix[
        ["Brand", "Product Name", "SKU", "Barcode Text"] + month_columns
    ].sort_values(["Brand", "Product Name", "SKU", "Barcode Text"])

    top_matrix_brands = grouped.head(25)["Brand"].astype(str).tolist()
    customer_brand_matrix = (
        sales[sales["Brand"].astype(str).isin(top_matrix_brands)]
        .pivot_table(index="Customer Name", columns="Brand", values="Line Total", aggfunc="sum", fill_value=0)
        .reset_index()
    )
    customer_brand_matrix = customer_brand_matrix.merge(
        customer_grouped[["Customer Name", "Revenue"]], on="Customer Name", how="left"
    ).sort_values("Revenue", ascending=False)
    customer_brand_matrix = customer_brand_matrix.drop(columns=["Revenue"])

    total_quantity = float(grouped["Quantity"].sum())
    total_revenue = float(grouped["Revenue"].sum())
    total_skus = int(sales.loc[sales["SKU"] != "", "SKU"].nunique())

    brand_cards = []
    for _, row in grouped.iterrows():
        share = row["Revenue"] / total_revenue if total_revenue else 0
        brand = str(row["Brand"])
        product_rows = []
        for _, product in product_grouped[product_grouped["Brand"].eq(brand)].iterrows():
            product_rows.append(
                "<tr>"
                f"<td>{escape(text(product['Product Name']))}</td>"
                f"<td>{escape(text(product['SKU']))}</td>"
                f"<td>{escape(text(product['Barcode Text']))}</td>"
                f"<td class='num'>{number(product['Quantity'])}</td>"
                f"<td class='num'>{money(product['Revenue'])}</td>"
                f"<td class='num'>{price_label(product['MinPrice'], product['MaxPrice'])}</td>"
                "</tr>"
            )
        brand_cards.append(
            '<details class="brand-detail">'
            "<summary>"
            f"<span class='brand-name'>{escape(brand)}</span>"
            f"<span class='num'>{number(row['Quantity'])}</span>"
            f"<span class='num'>{money(row['Revenue'])}</span>"
            f"<span class='num'>{int(row['SkuCount'])}</span>"
            f"<span class='num'>{pct(share)}</span>"
            "</summary>"
            "<div class='product-detail'>"
            "<table>"
            "<thead><tr><th>Product</th><th>SKU</th><th>Barcode</th><th class='num'>Units Sold</th><th class='num'>Sales</th><th class='num'>Price</th></tr></thead>"
            f"<tbody>{''.join(product_rows)}</tbody>"
            "</table>"
            "</div>"
            "</details>"
        )

    top_brands = grouped.head(8).copy()
    max_revenue = float(top_brands["Revenue"].max()) if len(top_brands) else 1.0
    bar_rows = []
    for _, row in top_brands.iterrows():
        width = row["Revenue"] / max_revenue * 100 if max_revenue else 0
        bar_rows.append(
            "<div class='bar-row'>"
            f"<div class='bar-label'>{escape(str(row['Brand']))}</div>"
            f"<div class='bar-track'><div class='bar-fill' style='width: {width:.1f}%'></div></div>"
            f"<div class='bar-value'>{money(row['Revenue'])}</div>"
            "</div>"
        )

    unmatched = sales[sales["Brand"].eq("Unknown / SKU not in product list")]
    if len(unmatched):
        unmatched_grouped = (
            unmatched.groupby(["SKU", "Product"], dropna=False)
            .agg(Quantity=("Quantity", "sum"), Revenue=("Line Total", "sum"))
            .reset_index()
            .sort_values("Revenue", ascending=False)
            .head(50)
        )
        unmatched_rows = [
            "<tr>"
            f"<td>{escape(text(row['SKU']))}</td>"
            f"<td>{escape(text(row['Product']))}</td>"
            f"<td class='num'>{number(row['Quantity'])}</td>"
            f"<td class='num'>{money(row['Revenue'])}</td>"
            "</tr>"
            for _, row in unmatched_grouped.iterrows()
        ]
        unmatched_section = f"""
        <section>
          <h2>Unmatched SKUs</h2>
          <p class="note">These sales rows could not be matched to a Brand in the product file. They are included under "Unknown / SKU not in product list" in the brand table.</p>
          <table>
            <thead><tr><th>SKU</th><th>Product</th><th class="num">Units</th><th class="num">Sales</th></tr></thead>
            <tbody>{''.join(unmatched_rows)}</tbody>
          </table>
        </section>
        """
    else:
        unmatched_section = """
        <section>
          <h2>SKU Match Check</h2>
          <p class="note">All sold SKUs were matched to a Brand in the product file.</p>
        </section>
        """

    excel_output_file = write_excel_report(
        grouped,
        product_grouped,
        customer_grouped,
        customer_detail_grouped,
        customer_brand_matrix,
        product_monthly_grouped,
        product_monthly_matrix,
        total_revenue,
        excel_output_file,
    )

    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>April Brand Sales Report</title>
  <style>
    :root {{
      color-scheme: light;
      --ink: #182026;
      --muted: #68727c;
      --line: #d9dee4;
      --panel: #f7f8fa;
      --accent: #0f766e;
      --accent-soft: #d8efeb;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, "Microsoft YaHei", sans-serif;
      color: var(--ink);
      background: #ffffff;
      line-height: 1.45;
    }}
    header {{
      padding: 36px 42px 22px;
      border-bottom: 1px solid var(--line);
    }}
    h1 {{
      margin: 0 0 8px;
      font-size: 30px;
      letter-spacing: 0;
    }}
    h2 {{
      margin: 0 0 14px;
      font-size: 18px;
    }}
    .subtitle {{ color: var(--muted); margin: 0; }}
    main {{ padding: 28px 42px 44px; }}
    .kpis {{
      display: grid;
      grid-template-columns: repeat(4, minmax(160px, 1fr));
      gap: 14px;
      margin-bottom: 28px;
    }}
    .kpi {{
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      background: var(--panel);
    }}
    .kpi .label {{
      color: var(--muted);
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: .06em;
    }}
    .kpi .value {{
      margin-top: 6px;
      font-size: 24px;
      font-weight: 700;
    }}
    section {{ margin-top: 30px; }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      padding: 10px 8px;
      text-align: left;
      vertical-align: top;
    }}
    th {{
      background: #eef2f4;
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: .04em;
      color: #3c4852;
    }}
    .num {{ text-align: right; white-space: nowrap; }}
    .bar-row {{
      display: grid;
      grid-template-columns: minmax(180px, 260px) 1fr minmax(110px, auto);
      gap: 12px;
      align-items: center;
      margin: 10px 0;
    }}
    .bar-label {{ font-weight: 600; overflow-wrap: anywhere; }}
    .bar-track {{
      height: 18px;
      border-radius: 6px;
      background: var(--accent-soft);
      overflow: hidden;
    }}
    .bar-fill {{
      height: 100%;
      background: var(--accent);
    }}
    .bar-value {{ text-align: right; color: var(--muted); }}
    .brand-list {{
      border-top: 1px solid var(--line);
    }}
    .brand-header,
    .brand-detail summary {{
      display: grid;
      grid-template-columns: minmax(180px, 2fr) minmax(100px, 1fr) minmax(120px, 1fr) minmax(90px, .8fr) minmax(90px, .8fr);
      gap: 12px;
      align-items: center;
    }}
    .brand-header {{
      padding: 10px 12px;
      background: #eef2f4;
      color: #3c4852;
      font-size: 12px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: .04em;
      border-bottom: 1px solid var(--line);
    }}
    .brand-detail {{
      border-bottom: 1px solid var(--line);
    }}
    .brand-detail summary {{
      cursor: pointer;
      list-style: none;
      padding: 12px;
    }}
    .brand-detail summary::-webkit-details-marker {{
      display: none;
    }}
    .brand-detail summary::before {{
      content: "+";
      color: var(--accent);
      font-weight: 700;
      margin-right: 8px;
    }}
    .brand-detail[open] summary::before {{
      content: "-";
    }}
    .brand-name {{
      font-weight: 700;
      overflow-wrap: anywhere;
    }}
    .product-detail {{
      padding: 0 12px 16px 28px;
      overflow-x: auto;
    }}
    .note {{
      margin: 0 0 12px;
      color: var(--muted);
    }}
    .source {{
      margin-top: 30px;
      color: var(--muted);
      font-size: 12px;
    }}
    @media (max-width: 860px) {{
      header, main {{ padding-left: 18px; padding-right: 18px; }}
      .kpis {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .bar-row {{ grid-template-columns: 1fr; gap: 5px; }}
      .bar-value {{ text-align: left; }}
      .brand-header {{ display: none; }}
      .brand-detail summary {{ grid-template-columns: 1fr; gap: 5px; }}
      .brand-detail summary .num {{ text-align: left; }}
      .product-detail {{ padding-left: 12px; }}
      table {{ font-size: 12px; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>April Brand Sales Report</h1>
  </header>
  <main>
    <div class="kpis">
      <div class="kpi"><div class="label">Total Sales</div><div class="value">{money(total_revenue)}</div></div>
      <div class="kpi"><div class="label">Units Sold</div><div class="value">{number(total_quantity)}</div></div>
      <div class="kpi"><div class="label">Sold SKUs</div><div class="value">{total_skus}</div></div>
      <div class="kpi"><div class="label">Brands</div><div class="value">{len(grouped)}</div></div>
    </div>

    <section>
      <h2>Top Brands by Sales</h2>
      {''.join(bar_rows)}
    </section>

    <section>
      <h2>Brand Detail</h2>
      <div class="brand-list">
        <div class="brand-header">
          <span>Brand</span>
          <span class="num">Units Sold</span>
          <span class="num">Sales</span>
          <span class="num">Sold SKUs</span>
          <span class="num">Sales Share</span>
        </div>
        {''.join(brand_cards)}
      </div>
    </section>

    {unmatched_section}

    <div class="source">
      Source files: {escape(SALES_FILE.name)} and {escape(PRODUCTS_FILE.name)}. Report generated on {datetime.now():%Y-%m-%d %H:%M}.
    </div>
  </main>
</body>
</html>
"""

    html = build_interactive_html_report(
        grouped,
        product_grouped,
        customer_grouped,
        customer_detail_grouped,
        customer_monthly_grouped,
        product_monthly_grouped,
        product_monthly_matrix,
        date_range,
        total_quantity,
        total_revenue,
        total_skus,
        report_title,
    )
    html_output_file.write_text(html, encoding="utf-8")

    print(f"OutputFile: {html_output_file}")
    print(f"ExcelOutputFile: {excel_output_file}")
    print(f"SalesRows: {len(sales)}")
    print(f"ProductRows: {len(products)}")
    print(f"Brands: {len(grouped)}")
    print(f"TotalUnits: {total_quantity:.0f}")
    print(f"TotalSales: {total_revenue:.2f}")
    print(f"UnmatchedSkus: {sales.loc[sales['Brand'].eq('Unknown / SKU not in product list'), 'SKU'].nunique()}")


if __name__ == "__main__":
    main()
